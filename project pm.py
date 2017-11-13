#!/usr/bin/python3
# -*- coding: utf-8 -*-
import openpyxl
import openpyxl.cell
from datetime import datetime, date, time
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Border, Side, Alignment
from openpyxl.styles import Font
import os



# Ввод значений для поиска (список)

def set_go_values():
    go_values = input('Пожалуйста, введите должностных лиц через запятую: ').split(sep=', ')
    print("Ваши данные приняты, пожалуйста подождите...")
    return go_values


search_list = set_go_values()


# вывод списка файлов
def scan_files():
    files = os.listdir(os.getcwd())
    true_files = list(filter(lambda x: x.endswith('.xlsx'), files))
    return (true_files)


file_list = scan_files()
data_file = dict()
file_name = 'СБ-ПМ-{} от {}.xlsx'.format('-'.join(search_list), str(datetime.date(datetime.now())))


# Выгрузка данных с выбранной строки
def get_data_string(row_begin, sheet1):
    data = list()
    max_col = sheet1.max_column
    for num_column in range(1, max_col, 1):
        data.append(sheet1.cell(row=row_begin, column=num_column).value)
    return data


# Выгрузка данных с файла по выбранным данным
def set_begin_data_rows(go_value, sheet):
    nums_rw = list()
    dt_fl = list()
    row_end = sheet.max_row
    for num_row in range(1, row_end, 1):
        st_sh = sheet.cell(row=num_row, column=3).value
        if type(st_sh) == type(str()):
            if go_value in st_sh:
                nums_rw.append(num_row)
    for num, rw in enumerate(nums_rw):
        dt_fl.append(get_data_string(rw, sheet))
    return dt_fl


def structurate_data():
    for f in file_list:
        wb = openpyxl.load_workbook(f)
        sheet = wb.get_active_sheet()
        l = list()
        for nm in search_list:
            print('Ищу данные по {} в файле {}.'.format(nm, f))
            data_file['ПМ {} из файла {}'.format(nm, f)] = set_begin_data_rows(nm, sheet)

def create_header(sheet_dt):
    sheet_dt.cell(row=1,column=1).value = 'Сборный ПМ-{} от {}'.format('-'.join(search_list), str(datetime.date(datetime.now())))
    sheet_dt.merge_cells('A1:H1')



def create_new_book():
    wb_rec = openpyxl.Workbook()
    sheet_rec = wb_rec.get_active_sheet()
    sheet_rec.title = 'ПМ'
    create_header(sheet_rec)
    r = 1
    c = 0
    for n, d in data_file.items():
        r = r + 1
        sheet_rec.cell(row=r + 1, column=1).value = n
        sheet_rec.merge_cells('A{}:H{}'.format((r+1),(r+1)))
        for i in d:
            r = r + 1
            c = 0
            for i1 in i:
                sheet_rec.cell(row=r+1, column=c+1).value = i1
                c = c + 1
    wb_rec.save(file_name)

def style_normalize():
    print('Запись данных в файл успешно завершена. Теперь необходимо навести красоту.')
    wb_dt = openpyxl.Workbook(file_name)
    sheet_dt = wb_dt.get_sheet_names('ПМ')
    sheet_dt.cell(row=1,column=1).value = 'Cборный ПМ'
    sheet_dt.merge_cells('A1:H1')

structurate_data()
create_new_book()
